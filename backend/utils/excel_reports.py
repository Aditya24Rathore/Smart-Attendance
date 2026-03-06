import os
import tempfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from models import db, Student, Subject, Attendance, ClassSession, User


def generate_attendance_excel(department=None, subject_id=None, date_from=None, date_to=None):
    """
    Generate Excel report with two sheets:
    Sheet 1: Student-wise attendance percentage by subject
    Sheet 2: Daily attendance records
    """
    wb = Workbook()

    # ---- Sheet 1: Attendance Percentage Summary ----
    ws1 = wb.active
    ws1.title = "Attendance Summary"

    # Styles
    header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    title_font = Font(name='Calibri', bold=True, size=14)
    good_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    warn_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    bad_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # Title
    ws1['A1'] = 'Smart Attendance System - Attendance Summary Report'
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:H1')
    ws1['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    if department:
        ws1['A3'] = f'Department: {department}'

    # Build student query
    student_query = Student.query.join(User)
    if department:
        student_query = student_query.filter(Student.department == department)
    students = student_query.order_by(Student.roll_number).all()

    # Build subject query
    subject_query = Subject.query
    if department:
        subject_query = subject_query.filter_by(department=department)
    if subject_id:
        subject_query = subject_query.filter_by(id=subject_id)
    subjects = subject_query.order_by(Subject.code).all()

    # Headers for Sheet 1
    row = 5
    headers = ['Roll Number', 'Student Name', 'Department', 'Semester']
    for subj in subjects:
        headers.append(f'{subj.code}\n({subj.name})')
    headers.append('Overall %')

    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Data rows
    for student in students:
        row += 1
        ws1.cell(row=row, column=1, value=student.roll_number).border = thin_border
        ws1.cell(row=row, column=2, value=student.user.full_name if student.user else '').border = thin_border
        ws1.cell(row=row, column=3, value=student.department).border = thin_border
        ws1.cell(row=row, column=4, value=student.semester).border = thin_border

        total_all = 0
        present_all = 0

        for i, subj in enumerate(subjects):
            session_query = ClassSession.query.filter_by(subject_id=subj.id, is_active=False)
            if date_from:
                session_query = session_query.filter(ClassSession.date >= date_from)
            if date_to:
                session_query = session_query.filter(ClassSession.date <= date_to)
            total_sessions = session_query.count()

            present_count = Attendance.query.filter_by(
                student_id=student.id, subject_id=subj.id
            ).filter(Attendance.status.in_(['present', 'late'])).count()

            total_all += total_sessions
            present_all += present_count

            pct = round((present_count / total_sessions * 100), 1) if total_sessions > 0 else 0
            cell = ws1.cell(row=row, column=5 + i, value=f'{pct}%')
            cell.alignment = center_align
            cell.border = thin_border

            if pct >= 75:
                cell.fill = good_fill
            elif pct >= 50:
                cell.fill = warn_fill
            else:
                cell.fill = bad_fill

        overall = round((present_all / total_all * 100), 1) if total_all > 0 else 0
        overall_cell = ws1.cell(row=row, column=5 + len(subjects), value=f'{overall}%')
        overall_cell.alignment = center_align
        overall_cell.border = thin_border
        overall_cell.font = Font(bold=True)
        if overall >= 75:
            overall_cell.fill = good_fill
        elif overall >= 50:
            overall_cell.fill = warn_fill
        else:
            overall_cell.fill = bad_fill

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws1.column_dimensions[chr(64 + col) if col <= 26 else 'A'].width = 18

    # ---- Sheet 2: Daily Attendance Records ----
    ws2 = wb.create_sheet("Daily Records")
    ws2['A1'] = 'Smart Attendance System - Daily Attendance Records'
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:G1')
    ws2['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'

    row2 = 4
    daily_headers = ['Date', 'Subject', 'Roll Number', 'Student Name', 'Status', 'Marked At', 'Marked By']
    for col, header in enumerate(daily_headers, 1):
        cell = ws2.cell(row=row2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Build attendance query
    att_query = Attendance.query.join(ClassSession).join(Student).join(Subject,
        Attendance.subject_id == Subject.id)
    if department:
        att_query = att_query.filter(ClassSession.department == department)
    if subject_id:
        att_query = att_query.filter(Attendance.subject_id == subject_id)
    if date_from:
        att_query = att_query.filter(ClassSession.date >= date_from)
    if date_to:
        att_query = att_query.filter(ClassSession.date <= date_to)

    records = att_query.order_by(ClassSession.date.desc(), Student.roll_number).all()

    for record in records:
        row2 += 1
        session_obj = record.session
        ws2.cell(row=row2, column=1,
                 value=session_obj.date.strftime('%Y-%m-%d') if session_obj.date else '').border = thin_border
        ws2.cell(row=row2, column=2,
                 value=record.subject.name if record.subject else '').border = thin_border
        ws2.cell(row=row2, column=3,
                 value=record.student.roll_number if record.student else '').border = thin_border
        ws2.cell(row=row2, column=4,
                 value=record.student.user.full_name if record.student and record.student.user else '').border = thin_border

        status_cell = ws2.cell(row=row2, column=5, value=record.status.upper())
        status_cell.alignment = center_align
        status_cell.border = thin_border
        if record.status == 'present':
            status_cell.fill = good_fill
        elif record.status == 'late':
            status_cell.fill = warn_fill
        else:
            status_cell.fill = bad_fill

        ws2.cell(row=row2, column=6,
                 value=record.marked_at.strftime('%H:%M:%S') if record.marked_at else '').border = thin_border
        ws2.cell(row=row2, column=7, value=record.marked_by or '').border = thin_border

    for col in range(1, 8):
        ws2.column_dimensions[chr(64 + col)].width = 20

    # Save to temp file
    temp_dir = tempfile.gettempdir()
    filepath = os.path.join(temp_dir, f'attendance_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    wb.save(filepath)
    return filepath
